from __future__ import annotations

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import BorderSide, CellStyle, Theme
from .utils import calc_col_widths


def _side(s: BorderSide) -> Side:
    return Side(border_style=s.style, color=s.color if s.style else "000000")


def _apply_cell_style(cell, style: CellStyle) -> None:
    cell.font = Font(
        name=style.font.name,
        size=style.font.size,
        bold=style.font.bold,
        italic=style.font.italic,
        color=style.font.color,
    )
    if style.fill.fg_color:
        cell.fill = PatternFill(
            fill_type=style.fill.pattern,
            fgColor=style.fill.fg_color,
        )
    b = style.border
    cell.border = Border(
        top=_side(b.top),
        bottom=_side(b.bottom),
        left=_side(b.left),
        right=_side(b.right),
    )
    cell.alignment = Alignment(
        horizontal=style.horizontal_align,
        wrap_text=style.wrap_text,
    )


def write(
    path: str,
    headers: list[str],
    rows: list[list],
    theme: Theme,
    column_mode: str | int | None,
    gradient_rules: list[dict],
    freeze_header: bool,
    sheet_name: str = "Sheet1",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        _apply_cell_style(cell, theme.header)

    # Data rows
    for row_idx, row in enumerate(rows):
        ws.append(row)
        style: CellStyle = theme.even_row if row_idx % 2 == 0 else theme.odd_row
        excel_row = row_idx + 2  # 1-indexed, row 1 is header
        for cell in ws[excel_row]:
            _apply_cell_style(cell, style)

    # Column widths
    if column_mode is not None:
        if column_mode == "fit" or column_mode == "auto":
            widths = calc_col_widths(headers, rows, padding=theme.column_padding)
            for col_idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        elif isinstance(column_mode, int):
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = column_mode

    # Color gradient rules (conditional formatting)
    for rule in gradient_rules:
        col_name = rule["column"]
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
        elif isinstance(col_name, int):
            col_idx = col_name + 1
        else:
            continue
        col_letter = get_column_letter(col_idx)
        data_rows = len(rows)
        if data_rows == 0:
            continue
        ref = f"{col_letter}2:{col_letter}{data_rows + 1}"
        low = rule["low"].lstrip("#")
        high = rule["high"].lstrip("#")
        ws.conditional_formatting.add(
            ref,
            ColorScaleRule(
                start_type="min",
                start_color=low,
                end_type="max",
                end_color=high,
            ),
        )

    # Freeze header
    if freeze_header:
        ws.freeze_panes = "A2"

    wb.save(path)
