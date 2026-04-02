import pytest

from xlutils import style
from xlutils.sheet_spec import SheetSpec
from xlutils.styler import Styler

DATA = [{"product": "A", "revenue": 100}, {"product": "B", "revenue": 200}]
DATA2 = [{"city": "Paris", "pop": 2000000}, {"city": "Lyon", "pop": 500000}]


# ---------------------------------------------------------------------------
# Single-sheet (primary) mode — existing API, backward-compat
# ---------------------------------------------------------------------------


def test_default_state():
    s = style(DATA)
    sheet = s._sheets[0]
    assert s._default_theme.name == "default"
    assert s._default_column_mode == "fit"
    assert s._default_freeze_header is True
    assert sheet._gradient_rules == []


def test_apply_theme_by_name():
    s = style(DATA).apply_theme("minimal")
    assert s._sheets[0]._theme.name == "minimal"


def test_apply_theme_object():
    from xlutils.themes import THEME_REGISTRY

    t = THEME_REGISTRY["minimal"]
    s = style(DATA).apply_theme(t)
    assert s._sheets[0]._theme is t


def test_unknown_theme_raises():
    with pytest.raises(ValueError, match="Unknown theme"):
        style(DATA).apply_theme("nonexistent")


def test_expand_columns():
    s = style(DATA).expand_columns(20)
    assert s._sheets[0]._column_mode == 20


def test_freeze_header():
    s = style(DATA).freeze_header(False)
    assert s._sheets[0]._freeze_header is False


def test_color_gradient_accumulates():
    s = (
        style(DATA)
        .color_gradient("revenue")
        .color_gradient("product", "#FF0000", "#00FF00")
    )
    assert len(s._sheets[0]._gradient_rules) == 2
    assert s._sheets[0]._gradient_rules[0]["column"] == "revenue"
    assert s._sheets[0]._gradient_rules[1]["low"] == "#FF0000"


def test_chain_returns_same_instance():
    s = style(DATA)
    assert s.apply_theme("minimal") is s
    assert s.expand_columns("fit") is s
    assert s.freeze_header() is s
    assert s.color_gradient("revenue") is s


# ---------------------------------------------------------------------------
# Multi-sheet mode — new API
# ---------------------------------------------------------------------------


def test_add_sheet_returns_sheet_spec():
    sp = style().add_sheet(DATA, "Sheet1")
    assert isinstance(sp, SheetSpec)


def test_sheet_spec_back_returns_styler():
    styler = style()
    sp = styler.add_sheet(DATA, "Sheet1")
    assert sp.back() is styler


def test_sheet_spec_save_delegates_to_styler(tmp_path):
    path = str(tmp_path / "out.xlsx")
    style().add_sheet(DATA, "Sheet1").save(path)
    import os
    assert os.path.exists(path)


def test_workbook_level_theme_inherited_by_all_sheets():
    s = style().apply_theme("minimal")
    s.add_sheet(DATA, "A").back()
    s.add_sheet(DATA2, "B").back()
    defaults = s._defaults()
    assert s._sheets[0]._resolve(defaults)["theme"].name == "minimal"
    assert s._sheets[1]._resolve(defaults)["theme"].name == "minimal"


def test_per_sheet_theme_overrides_workbook_default():
    s = style().apply_theme("minimal")
    s.add_sheet(DATA, "A").apply_theme("default").back()
    s.add_sheet(DATA2, "B").back()
    defaults = s._defaults()
    assert s._sheets[0]._resolve(defaults)["theme"].name == "default"
    assert s._sheets[1]._resolve(defaults)["theme"].name == "minimal"


def test_color_gradient_on_styler_in_multi_sheet_mode_raises():
    with pytest.raises(TypeError, match="color_gradient"):
        style().add_sheet(DATA, "A").back().color_gradient("revenue")


def test_backward_compat_single_sheet_style_with_data(tmp_path):
    from openpyxl import load_workbook
    path = str(tmp_path / "out.xlsx")
    style(DATA).apply_theme("minimal").save(path)
    wb = load_workbook(path)
    assert wb.active.title == "Sheet1"
    assert wb.active["A1"].font.bold is True


def test_backward_compat_sheet_name_kwarg(tmp_path):
    from openpyxl import load_workbook
    path = str(tmp_path / "out.xlsx")
    style(DATA).save(path, sheet_name="MySheet")
    wb = load_workbook(path)
    assert "MySheet" in wb.sheetnames


def test_duplicate_sheet_name_raises():
    with pytest.raises(ValueError, match="Duplicate sheet name"):
        style().add_sheet(DATA, "Sales").back().add_sheet(DATA2, "Sales")


def test_save_with_no_sheets_raises(tmp_path):
    with pytest.raises(ValueError, match="no sheets"):
        style().save(str(tmp_path / "out.xlsx"))


def test_mixed_mode_raises():
    with pytest.raises(TypeError, match="style\\(data\\)"):
        style(DATA).add_sheet(DATA2, "Extra")


def test_sheet_name_kwarg_with_multiple_sheets_raises(tmp_path):
    with pytest.raises(ValueError, match="sheet_name kwarg"):
        (style()
            .add_sheet(DATA, "A").back()
            .add_sheet(DATA2, "B").back()
            .save(str(tmp_path / "out.xlsx"), sheet_name="X"))


def test_sheet_spec_chain_returns_same_spec():
    sp = style().add_sheet(DATA, "A")
    assert sp.apply_theme("minimal") is sp
    assert sp.expand_columns(20) is sp
    assert sp.freeze_header(False) is sp
    assert sp.color_gradient("revenue") is sp


def test_multi_sheet_no_data_styler_is_not_primary():
    s = style()
    assert s._primary is False
    assert s._sheets == []
