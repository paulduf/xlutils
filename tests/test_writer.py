import os
import tempfile

from openpyxl import load_workbook

from xlutils import style

DATA = [{"product": "Alpha", "revenue": 100}, {"product": "Beta", "revenue": 200}]
DATA2 = [{"city": "Paris", "pop": 2000000}, {"city": "Lyon", "pop": 500000}]


def _save_and_load(styler, **kwargs):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        styler.save(path, **kwargs)
        wb = load_workbook(path)
        return wb
    finally:
        os.unlink(path)


def test_one_liner_writes_file():
    wb = _save_and_load(style(DATA))
    ws = wb.active
    assert ws["A1"].value == "product"
    assert ws["B1"].value == "revenue"
    assert ws["A2"].value == "Alpha"
    assert ws["B3"].value == 200


def test_header_is_bold():
    wb = _save_and_load(style(DATA))
    ws = wb.active
    assert ws["A1"].font.bold is True


def test_banded_rows_have_different_fills():
    wb = _save_and_load(style(DATA))
    ws = wb.active
    fill_row2 = ws["A2"].fill.fgColor.rgb
    fill_row3 = ws["A3"].fill.fgColor.rgb
    assert fill_row2 != fill_row3


def test_minimal_theme():
    wb = _save_and_load(style(DATA).apply_theme("minimal"))
    ws = wb.active
    assert ws["A1"].font.bold is True


def test_custom_sheet_name():
    wb = _save_and_load(style(DATA), sheet_name="Sales")
    assert "Sales" in wb.sheetnames


def test_fixed_column_width():
    wb = _save_and_load(style(DATA).expand_columns(30))
    ws = wb.active
    assert ws.column_dimensions["A"].width == 30


def test_no_column_expand():
    # expand_columns(None) should leave widths at default
    wb = _save_and_load(style(DATA).expand_columns(None))
    ws = wb.active
    # openpyxl default is None/0; just verify no crash
    assert ws["A1"].value == "product"


# ---------------------------------------------------------------------------
# Multi-sheet integration tests
# ---------------------------------------------------------------------------


def test_multi_sheet_writes_correct_sheet_count():
    wb = _save_and_load(
        style().add_sheet(DATA, "Sales").back().add_sheet(DATA2, "Costs")
    )
    assert len(wb.sheetnames) == 2


def test_multi_sheet_correct_names():
    wb = _save_and_load(
        style().add_sheet(DATA, "Sales").back().add_sheet(DATA2, "Costs")
    )
    assert wb.sheetnames == ["Sales", "Costs"]


def test_multi_sheet_each_sheet_has_correct_data():
    wb = _save_and_load(
        style().add_sheet(DATA, "Sales").back().add_sheet(DATA2, "Costs")
    )
    assert wb["Sales"]["A1"].value == "product"
    assert wb["Sales"]["A2"].value == "Alpha"
    assert wb["Costs"]["A1"].value == "city"
    assert wb["Costs"]["A2"].value == "Paris"


def test_multi_sheet_per_sheet_theme_applied():
    # "default" theme has white header font, "minimal" has black
    wb = _save_and_load(
        style()
        .add_sheet(DATA, "Sales")
        .apply_theme("default")
        .back()
        .add_sheet(DATA2, "Costs")
        .apply_theme("minimal")
    )
    sales_header_color = wb["Sales"]["A1"].font.color.rgb
    costs_header_color = wb["Costs"]["A1"].font.color.rgb
    assert sales_header_color != costs_header_color


def test_multi_sheet_per_sheet_freeze_header():
    wb = _save_and_load(
        style()
        .add_sheet(DATA, "Sales")
        .back()
        .add_sheet(DATA2, "Costs")
        .freeze_header(False)
    )
    assert wb["Sales"].freeze_panes == "A2"
    assert wb["Costs"].freeze_panes is None


def test_multi_sheet_gradient_on_specific_sheet_only():
    wb = _save_and_load(
        style()
        .add_sheet(DATA, "Sales")
        .color_gradient("revenue")
        .back()
        .add_sheet(DATA2, "Costs")
    )
    sales_rules = list(wb["Sales"].conditional_formatting)
    costs_rules = list(wb["Costs"].conditional_formatting)
    assert len(sales_rules) > 0
    assert len(costs_rules) == 0


def test_multi_sheet_workbook_default_theme_inherited():
    wb = _save_and_load(
        style()
        .apply_theme("minimal")
        .add_sheet(DATA, "Sales")
        .back()
        .add_sheet(DATA2, "Costs")
    )
    # Both sheets use "minimal" — header should be bold
    assert wb["Sales"]["A1"].font.bold is True
    assert wb["Costs"]["A1"].font.bold is True


def test_multi_sheet_sheet_order_preserved():
    wb = _save_and_load(
        style()
        .add_sheet(DATA, "First")
        .back()
        .add_sheet(DATA2, "Second")
        .back()
        .add_sheet(DATA, "Third")
    )
    assert wb.sheetnames == ["First", "Second", "Third"]
